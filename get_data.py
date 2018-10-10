# 0.0 coding:utf-8 0.0
#获取数据
from openpyxl import load_workbook

def get_data(filename,filecolum):
    book = load_workbook(filename)  #打开excel文件
    sheet = book.get_sheet_by_name('Sheet1')   #读取相应的sheet
    row_num =1
    file_data = []  #数据存储数组
    while row_num <98:
        file_data.append(sheet.cell(row=row_num,column=filecolum).value)
        row_num=row_num+1
    return file_data

#filename1 = '201.xlsx'
#filecolum1 = 2
#d = get_data(filename1,filecolum1)
#print(d)